#!/usr/bin/python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from typing import List, Dict, Any
import io
import re
from types import SimpleNamespace
from sqlalchemy import select, false
from xlsxwriter.utility import xl_col_to_name
from apps.vadmin.sport.models import (
    VadminPefClass,
    VadminPefGrade,
    VadminPefSchool,
    VadminPefStudent,
    VadminSportBatch,
    VadminSportStandardItem
)
from apps.vadmin.sport.service.rule_engine import RuleEngine
from apps.vadmin.sport.service.scope_service import match_scope_by_name

class BatchImportService:
    SCORE_HEADERS = [
        ("学号", "student_no"),
        ("姓名", "student_name"),
        ("性别", "gender"),
        ("学校", "school_name"),
        ("年级", "grade_name"),
        ("班级", "class_name"),
        ("项目", "item_name"),
        ("成绩", "raw_score")
    ]
    SCORE_REQUIRED_LABELS = {label for label, _ in SCORE_HEADERS}
    SCORE_WIDE_HEADERS = [
        ("\u8eab\u4efd\u8bc1", "id_card"),
        ("\u5b66\u751f\u59d3\u540d", "student_name"),
        ("\u6027\u522b", "gender"),
        ("\u5b66\u6821", "school_name"),
        ("\u5e74\u7ea7", "grade_name"),
        ("\u73ed\u7ea7", "class_name"),
        ("\u8eab\u9ad8", "height"),
        ("\u4f53\u91cd", "weight"),
        ("BMI", "bmi"),
        ("\u80ba\u6d3b\u91cf", "lung"),
        ("50\u7c73\u8dd1", "run_50"),
        ("\u5750\u4f4d\u4f53\u524d\u5c48", "sit"),
        ("\u7acb\u5b9a\u8df3\u8fdc", "jump"),
        ("\u5f15\u4f53\u5411\u4e0a", "pull_up"),
        ("1\u5206\u949f\u4ef0\u5367\u8d77\u5750", "situp"),
        ("1000\u7c73\u8dd1", "run_1000"),
        ("800\u7c73\u8dd1", "run_800"),
        ("\u8001\u5e08\u8bc4\u8bed", "teacher_comment")
    ]
    SCORE_WIDE_ITEM_HEADERS = [
        ("\u8eab\u9ad8", "height"),
        ("\u4f53\u91cd", "weight"),
        ("BMI", "bmi"),
        ("\u80ba\u6d3b\u91cf", "lung"),
        ("50\u7c73\u8dd1", "run_50"),
        ("\u5750\u4f4d\u4f53\u524d\u5c48", "sit"),
        ("\u7acb\u5b9a\u8df3\u8fdc", "jump"),
        ("\u5f15\u4f53\u5411\u4e0a", "pull_up"),
        ("1\u5206\u949f\u4ef0\u5367\u8d77\u5750", "situp"),
        ("1000\u7c73\u8dd1", "run_1000"),
        ("800\u7c73\u8dd1", "run_800")
    ]
    SCORE_WIDE_BASE_REQUIRED_LABELS = {"\u8eab\u4efd\u8bc1", "\u5b66\u751f\u59d3\u540d", "\u6027\u522b", "\u5b66\u6821", "\u5e74\u7ea7", "\u73ed\u7ea7"}
    UNLIMITED_VALUES = {'', '*', 'all', '不限', '全局', '全部'}
    FITNESS_REMOVED_ITEM_CODES = {'run_50x8'}
    FITNESS_REMOVED_ITEM_NAMES = {'50米x8', '50米-8', '50x8'}
    FITNESS_VIRTUAL_ITEMS = [
        {"item_code": "height", "item_name": "身高", "gender": "all", "calc_mode": "record", "sort": -2},
        {"item_code": "weight", "item_name": "体重", "gender": "all", "calc_mode": "record", "sort": -1}
    ]

    @staticmethod
    def _clean_cell(value) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        if text.endswith('.0') and text[:-2].isdigit():
            return text[:-2]
        return text

    @staticmethod
    def _normalize_header(value) -> str:
        text = BatchImportService._clean_cell(value)
        return text.replace('*', '').replace(' ', '').replace('\u3000', '')

    @staticmethod
    def _normalize_item_key(value) -> str:
        text = BatchImportService._clean_cell(value)
        if not text:
            return ''
        text = text.lower()
        text = re.sub(r'[\\s_*\u3000\\(\\)\\-\\u00b7\\u2022,，。/:;!?%]', '', text)
        text = text.replace('\u4e00\u5206\u949f', '1')
        return text

    @staticmethod
    def _build_item_aliases(item_name: str, item_code: str | None = None) -> list[str]:
        aliases: set[str] = set()
        norm_name = BatchImportService._normalize_item_key(item_name)
        if norm_name:
            aliases.add(norm_name)
        code = BatchImportService._normalize_item_key(item_code) if item_code else ''
        if code:
            aliases.add(code)
            if code == 'height':
                aliases.update({'height', '身高'})
            elif code == 'weight':
                aliases.update({'weight', '体重'})
            elif code == 'bmi':
                aliases.add('bmi')
            elif code in {'lung', 'vitalcapacity', 'pulmonary'}:
                aliases.update({'lung', 'pulmonary', '肺活量'})
            elif code in {'run50', 'run50m', '50'}:
                aliases.update({'50', '50米', '50米跑', 'run50'})
            elif code in {'run1000', 'run1000m', '1000'}:
                aliases.update({'1000', '1000米', '1000米跑', 'run1000'})
            elif code in {'run800', 'run800m', '800'}:
                aliases.update({'800', '800米', '800米跑', 'run800'})
            elif code in {'sit', 'sitandreach'}:
                aliases.update({'sit', 'sitandreach', '坐位体前屈', '坐位体前弯'})
            elif code in {'jump', 'standinglongjump'}:
                aliases.update({'jump', '立定跳远'})
            elif code in {'pullup', 'pullupcount'}:
                aliases.update({'pullup', '引体向上'})
            elif code in {'situp', 'situpcount'}:
                aliases.update({'situp', '仰卧起坐', '1分钟仰卧起坐', '一分钟仰卧起坐'})

        if norm_name:
            if '50' in norm_name or '50米' in item_name:
                aliases.update({'50', '50米', '50米跑', 'run50'})
            if '1000' in norm_name or '1000米' in item_name:
                aliases.update({'1000', '1000米', '1000米跑', 'run1000'})
            if '800' in norm_name or '800米' in item_name:
                aliases.update({'800', '800米', '800米跑', 'run800'})
            if '坐位体前屈' in item_name:
                aliases.update({'sitandreach', '坐位体前屈', '坐位体前弯'})
            if '坐位体前弯' in item_name:
                aliases.update({'sitandreach', '坐位体前屈', '坐位体前弯'})
            if '立定跳远' in item_name:
                aliases.update({'jump', '立定跳远'})
            if '引体向上' in item_name:
                aliases.update({'pullup', '引体向上'})
            if '仰卧起坐' in item_name:
                aliases.update({'situp', '仰卧起坐', '1分钟仰卧起坐', '一分钟仰卧起坐'})

        return list({BatchImportService._normalize_item_key(alias) for alias in aliases})

    @staticmethod
    def _normalize_item_name_for_lookup(value: str) -> str:
        return BatchImportService._normalize_item_key(value)

    @staticmethod
    def _is_unlimited(value) -> bool:
        return BatchImportService._clean_cell(value).lower() in BatchImportService.UNLIMITED_VALUES

    @staticmethod
    def _normalize_gender(value) -> str:
        text = BatchImportService._clean_cell(value).lower()
        if text in {'male', 'm', '1', '男', 'boy', 'man'}:
            return 'male'
        if text in {'female', 'f', '0', '2', '女', 'girl', 'woman'}:
            return 'female'
        return 'all'

    @staticmethod
    def _option_list(values: list[str]) -> list[dict[str, str]]:
        result: list[dict[str, str]] = []
        seen: set[str] = set()
        for value in values:
            text = BatchImportService._clean_cell(value)
            if text and text not in seen:
                seen.add(text)
                result.append({"label": text, "value": text})
        return result

    @staticmethod
    def is_removed_fitness_item(item) -> bool:
        code = BatchImportService._clean_cell(getattr(item, "item_code", "")).lower()
        name = BatchImportService._clean_cell(getattr(item, "item_name", ""))
        return code in BatchImportService.FITNESS_REMOVED_ITEM_CODES or name in BatchImportService.FITNESS_REMOVED_ITEM_NAMES

    @staticmethod
    def _allows_negative_score_item(item) -> bool:
        code = BatchImportService._clean_cell(getattr(item, "item_code", "")).lower()
        name = BatchImportService._clean_cell(getattr(item, "item_name", ""))
        return code == "sit" or "坐位体前屈" in name

    @staticmethod
    def normalize_standard_items(biz_type: str, standard_id: int | None, items: list) -> list:
        normalized = list(items or [])
        if biz_type != 'fitness':
            return normalized

        normalized = [item for item in normalized if not BatchImportService.is_removed_fitness_item(item)]
        existing_codes = {
            BatchImportService._clean_cell(getattr(item, "item_code", "")).lower()
            for item in normalized
        }
        for virtual_item in BatchImportService.FITNESS_VIRTUAL_ITEMS:
            if virtual_item["item_code"] in existing_codes:
                continue
            normalized.append(SimpleNamespace(
                standard_id=standard_id,
                item_code=virtual_item["item_code"],
                item_name=virtual_item["item_name"],
                gender=virtual_item["gender"],
                calc_mode=virtual_item["calc_mode"],
                pass_threshold=None,
                excellent_threshold=None,
                full_threshold=None,
                segment_json=[],
                is_required=True,
                is_gate_item=False,
                max_score=0,
                sort=virtual_item["sort"]
            ))
        return sorted(normalized, key=lambda item: (getattr(item, "sort", 0) or 0, getattr(item, "id", 0) or 0))

    @staticmethod
    def _score_template_headers(template_style: str = 'compact') -> list[dict]:
        template_headers = BatchImportService.SCORE_WIDE_HEADERS if template_style == 'wide' else BatchImportService.SCORE_HEADERS
        return [
            {
                "label": label,
                "field": field,
                "required": label in BatchImportService.SCORE_WIDE_BASE_REQUIRED_LABELS if template_style == 'wide' else True
            }
            for label, field in template_headers
        ]

    @staticmethod
    def format_score_errors(errors: list[str]) -> str:
        if not errors:
            return ""
        deduped_errors = list(dict.fromkeys(errors))
        visible_errors = deduped_errors[:30]
        suffix = f"，另有 {len(deduped_errors) - len(visible_errors)} 条错误未展示" if len(deduped_errors) > len(visible_errors) else ""
        return f"共 {len(visible_errors)} 条错误：{'； '.join(visible_errors)}{suffix}"

    @staticmethod
    def _in_batch_scope(batch: VadminSportBatch, school_name: str, grade_name: str, class_name: str) -> bool:
        return (
            (BatchImportService._is_unlimited(batch.school_name) or batch.school_name == school_name)
            and (BatchImportService._is_unlimited(batch.grade_name) or batch.grade_name == grade_name)
            and (BatchImportService._is_unlimited(batch.class_name) or batch.class_name == class_name)
        )

    @staticmethod
    async def _load_scope_rows(db, auth, batch: VadminSportBatch) -> list[tuple[str, str, str]]:
        rows = (await db.execute(
            select(VadminPefSchool.school_name, VadminPefGrade.grade_name, VadminPefClass.class_name)
            .select_from(VadminPefClass)
            .join(VadminPefGrade, VadminPefClass.grade_id == VadminPefGrade.id)
            .join(VadminPefSchool, VadminPefClass.school_id == VadminPefSchool.id)
            .where(
                VadminPefSchool.is_delete == false(),
                VadminPefGrade.is_delete == false(),
                VadminPefClass.is_delete == false(),
                VadminPefSchool.is_active == True,
                VadminPefGrade.is_active == True,
                VadminPefClass.is_active == True
            )
            .order_by(VadminPefSchool.sort.asc(), VadminPefGrade.sort.asc(), VadminPefClass.sort.asc())
        )).all()
        result: list[tuple[str, str, str]] = []
        for school_name, grade_name, class_name in rows:
            school = BatchImportService._clean_cell(school_name)
            grade = BatchImportService._clean_cell(grade_name)
            cls = BatchImportService._clean_cell(class_name)
            if not BatchImportService._in_batch_scope(batch, school, grade, cls):
                continue
            if not match_scope_by_name(auth, school, cls):
                continue
            result.append((school, grade, cls))
        return result

    @staticmethod
    def _build_scope_option_maps(scope_rows: list[tuple[str, str, str]]) -> dict[str, Any]:
        schools: list[str] = []
        grades_by_school: dict[str, list[str]] = {}
        classes_by_school_grade: dict[str, list[str]] = {}
        for school, grade, cls in scope_rows:
            if school not in schools:
                schools.append(school)
            grades_by_school.setdefault(school, [])
            if grade not in grades_by_school[school]:
                grades_by_school[school].append(grade)
            key = f"{school}|{grade}"
            classes_by_school_grade.setdefault(key, [])
            if cls not in classes_by_school_grade[key]:
                classes_by_school_grade[key].append(cls)
        return {
            "schools": schools,
            "grades_by_school": grades_by_school,
            "classes_by_school_grade": classes_by_school_grade
        }

    @staticmethod
    def apply_score_template_validations(writer, options: dict[str, Any], max_row: int = 1000, template_style: str = 'compact') -> None:
        if not writer or not writer.wb or not writer.sheet:
            return
        wb = writer.wb
        sheet = writer.sheet
        option_sheet = wb.add_worksheet("_score_options")
        option_sheet.hide()
        template_headers = BatchImportService._score_template_headers(template_style)
        field_to_col = {header["field"]: index for index, header in enumerate(template_headers)}

        def write_column(col: int, values: list[str]) -> str | None:
            clean_values: list[str] = []
            seen_values: set[str] = set()
            for item in values:
                value = BatchImportService._clean_cell(item)
                if value and value not in seen_values:
                    seen_values.add(value)
                    clean_values.append(value)
            for row_idx, value in enumerate(clean_values):
                option_sheet.write(row_idx, col, value)
            if not clean_values:
                return None
            col_name = xl_col_to_name(col)
            return f"='_score_options'!${col_name}$1:${col_name}${len(clean_values)}"

        school_source = None
        if template_style == 'wide':
            id_card_source = write_column(0, options.get("student_ids") or options.get("id_cards") or [])
            student_name_source = write_column(1, options.get("student_names") or [])
            school_source = write_column(3, options.get("schools") or [])
            gender_source = write_column(2, ["男", "女"])
        else:
            school_source = write_column(field_to_col.get("school_name", 3), options.get("schools") or [])
            id_card_source = None
            student_no_source = write_column(field_to_col.get("student_no", 0), options.get("student_nos") or [])
            student_name_source = write_column(field_to_col.get("student_name", 1), options.get("student_names") or [])
            gender_source = write_column(field_to_col.get("gender", 2), ["男", "女"])
        item_name_source = write_column(field_to_col.get("item_name", 6), options.get("item_names") or [])

        school_col = field_to_col.get("school_name", 3)
        if school_source:
            sheet.data_validation(1, school_col, max_row, school_col, {'validate': 'list', 'source': school_source})

        if template_style == 'wide':
            if id_card_source:
                id_col = field_to_col.get("id_card", 0)
                sheet.data_validation(1, id_col, max_row, id_col, {'validate': 'list', 'source': id_card_source})
            sname_col = field_to_col.get("student_name", 1)
            if student_name_source:
                sheet.data_validation(1, sname_col, max_row, sname_col, {'validate': 'list', 'source': student_name_source})
            g_col = field_to_col.get("gender", 2)
            if gender_source:
                sheet.data_validation(1, g_col, max_row, g_col, {'validate': 'list', 'source': gender_source})
        else:
            no_col = field_to_col.get("student_no", 0)
            if student_no_source:
                sheet.data_validation(1, no_col, max_row, no_col, {'validate': 'list', 'source': student_no_source})
            sname_col = field_to_col.get("student_name", 1)
            if student_name_source:
                sheet.data_validation(1, sname_col, max_row, sname_col, {'validate': 'list', 'source': student_name_source})
            g_col = field_to_col.get("gender", 2)
            if gender_source:
                sheet.data_validation(1, g_col, max_row, g_col, {'validate': 'list', 'source': gender_source})
            item_col = field_to_col.get("item_name", 6)
            if item_name_source:
                sheet.data_validation(1, item_col, max_row, item_col, {'validate': 'list', 'source': item_name_source})

        grade_start_col = 10
        schools = options.get("schools") or []
        grades_by_school = options.get("grades_by_school") or {}
        for index, school in enumerate(schools):
            col = grade_start_col + index
            option_sheet.write(0, col, school)
            for row_idx, grade in enumerate(grades_by_school.get(school) or [], start=1):
                option_sheet.write(row_idx, col, grade)
        if schools:
            start_col = xl_col_to_name(grade_start_col)
            end_col = xl_col_to_name(grade_start_col + len(schools) - 1)
            grade_formula = (
                f"=OFFSET('_score_options'!${start_col}$2,0,"
                f"MATCH($D2,'_score_options'!${start_col}$1:${end_col}$1,0)-1,"
                f"COUNTA(OFFSET('_score_options'!${start_col}$2:${start_col}$1000,0,"
                f"MATCH($D2,'_score_options'!${start_col}$1:${end_col}$1,0)-1)),1)"
            )
            sheet.data_validation(1, 4, max_row, 4, {'validate': 'list', 'source': grade_formula})

        class_start_col = 80
        class_keys = list((options.get("classes_by_school_grade") or {}).keys())
        classes_by_school_grade = options.get("classes_by_school_grade") or {}
        for index, key in enumerate(class_keys):
            col = class_start_col + index
            option_sheet.write(0, col, key)
            for row_idx, cls in enumerate(classes_by_school_grade.get(key) or [], start=1):
                option_sheet.write(row_idx, col, cls)
        if class_keys:
            start_col = xl_col_to_name(class_start_col)
            end_col = xl_col_to_name(class_start_col + len(class_keys) - 1)
            class_formula = (
                f"=OFFSET('_score_options'!${start_col}$2,0,"
                f"MATCH($D2&\"|\"&$E2,'_score_options'!${start_col}$1:${end_col}$1,0)-1,"
                f"COUNTA(OFFSET('_score_options'!${start_col}$2:${start_col}$1000,0,"
                f"MATCH($D2&\"|\"&$E2,'_score_options'!${start_col}$1:${end_col}$1,0)-1)),1)"
            )
            sheet.data_validation(1, 5, max_row, 5, {'validate': 'list', 'source': class_formula})

    @staticmethod
    async def _load_batch(db, auth, biz_type: str, batch_id: int | None) -> VadminSportBatch:
        if not batch_id:
            raise ValueError("batch_id不能为空")
        batch = await db.scalar(select(VadminSportBatch).where(
            VadminSportBatch.id == batch_id,
            VadminSportBatch.biz_type == biz_type,
            VadminSportBatch.is_delete == false()
        ))
        if not batch:
            raise ValueError("批次不存在")
        if (
            not (BatchImportService._is_unlimited(batch.school_name) and BatchImportService._is_unlimited(batch.class_name))
            and not match_scope_by_name(auth, batch.school_name, batch.class_name)
        ):
            raise ValueError("无权访问该批次")
        return batch

    @staticmethod
    async def build_score_template_config(db, auth, biz_type: str, batch_id: int | None, template_style: str = 'compact') -> tuple[list[dict], dict[str, Any]]:
        batch = await BatchImportService._load_batch(db, auth, biz_type, batch_id)
        headers = BatchImportService._score_template_headers(template_style)
        scope_rows = await BatchImportService._load_scope_rows(db, auth, batch)
        scope_options = BatchImportService._build_scope_option_maps(scope_rows)
        student_rows = (await db.execute(
            select(VadminPefStudent, VadminPefSchool.school_name, VadminPefGrade.grade_name, VadminPefClass.class_name)
            .select_from(VadminPefStudent)
            .join(VadminPefSchool, VadminPefStudent.school_id == VadminPefSchool.id)
            .join(VadminPefGrade, VadminPefStudent.grade_id == VadminPefGrade.id)
            .join(VadminPefClass, VadminPefStudent.class_id == VadminPefClass.id)
            .where(
                VadminPefStudent.is_delete == false(),
                VadminPefSchool.is_delete == false(),
                VadminPefGrade.is_delete == false(),
                VadminPefClass.is_delete == false()
            )
            .order_by(VadminPefStudent.student_no.asc())
        )).all()
        standard_items = (await db.scalars(select(VadminSportStandardItem).where(
            VadminSportStandardItem.standard_id == batch.standard_id,
            VadminSportStandardItem.is_delete == false()
        ).order_by(VadminSportStandardItem.sort.asc(), VadminSportStandardItem.id.asc()))).all()
        standard_items = BatchImportService.normalize_standard_items(biz_type, batch.standard_id, standard_items)

        scoped_student_rows = [
            row for row in student_rows
            if BatchImportService._in_batch_scope(batch, row[1], row[2], row[3])
            and match_scope_by_name(auth, row[1], row[3])
        ]
        students = [row[0] for row in scoped_student_rows]
        options = {
            **scope_options,
            "student_nos": [item.student_no for item in students],
            "student_ids": [item.id_card for item in students],
            "id_cards": [item.id_card for item in students],
            "student_names": [item.name for item in students],
            "item_names": [item.item_name for item in standard_items]
        }
        return headers, options

    @staticmethod
    async def build_score_template_headers(db, auth, biz_type: str, batch_id: int | None, template_style: str = 'compact') -> list[dict]:
        headers, _ = await BatchImportService.build_score_template_config(db, auth, biz_type, batch_id, template_style)
        return headers

    @staticmethod
    async def validate_score_rows(db, auth, biz_type: str, batch_id: int | None, scores: list[dict]) -> tuple[VadminSportBatch, list[dict], list[str]]:
        batch = await BatchImportService._load_batch(db, auth, biz_type, batch_id)
        student_rows = (await db.execute(
            select(VadminPefStudent, VadminPefSchool.school_name, VadminPefGrade.grade_name, VadminPefClass.class_name)
            .select_from(VadminPefStudent)
            .join(VadminPefSchool, VadminPefStudent.school_id == VadminPefSchool.id)
            .join(VadminPefGrade, VadminPefStudent.grade_id == VadminPefGrade.id)
            .join(VadminPefClass, VadminPefStudent.class_id == VadminPefClass.id)
            .where(
                VadminPefStudent.is_delete == false(),
                VadminPefSchool.is_delete == false(),
                VadminPefGrade.is_delete == false(),
                VadminPefClass.is_delete == false()
            )
        )).all()
        student_map_by_no: dict[str, tuple] = {}
        student_map_by_id: dict[str, tuple] = {}
        for row in student_rows:
            student = row[0]
            student_no = BatchImportService._clean_cell(student.student_no)
            if student_no:
                student_map_by_no[student_no] = row
            id_card = BatchImportService._clean_cell(student.id_card)
            if id_card:
                student_map_by_id[id_card] = row

        standard_items = (await db.scalars(select(VadminSportStandardItem).where(
            VadminSportStandardItem.standard_id == batch.standard_id,
            VadminSportStandardItem.is_delete == false()
        ))).all()
        standard_items = BatchImportService.normalize_standard_items(biz_type, batch.standard_id, standard_items)
        item_name_map: dict[str, list[VadminSportStandardItem]] = {}
        for item in standard_items:
            for alias in BatchImportService._build_item_aliases(item.item_name, item.item_code):
                item_name_map.setdefault(alias, []).append(item)

        errors: list[str] = []
        normalized_scores: list[dict] = []
        for index, score in enumerate(scores, start=1):
            row_number = score.get("_row_number") or (index + 1)
            row_errors: list[str] = []
            student_no = BatchImportService._clean_cell(score.get("student_no"))
            id_card = BatchImportService._clean_cell(score.get("id_card"))
            student_name = BatchImportService._clean_cell(score.get("student_name"))
            gender = BatchImportService._clean_cell(score.get("gender"))
            school_name = BatchImportService._clean_cell(score.get("school_name"))
            grade_name = BatchImportService._clean_cell(score.get("grade_name"))
            class_name = BatchImportService._clean_cell(score.get("class_name"))
            item_name = BatchImportService._clean_cell(score.get("item_name"))
            raw_score = score.get("raw_score")

            if not BatchImportService._is_unlimited(batch.school_name) and school_name != batch.school_name:
                row_errors.append(f"school must match batch: {batch.school_name}")
            if not BatchImportService._is_unlimited(batch.grade_name) and grade_name != batch.grade_name:
                row_errors.append(f"grade must match batch: {batch.grade_name}")
            if not BatchImportService._is_unlimited(batch.class_name) and class_name != batch.class_name:
                row_errors.append(f"class must match batch: {batch.class_name}")
            if not match_scope_by_name(auth, school_name, class_name):
                row_errors.append("no permission to import this school/class")

            student_row = student_map_by_id.get(id_card) if id_card else None
            if not student_row and student_no:
                student_row = student_map_by_no.get(student_no)
            if not student_row and id_card:
                student_row = student_map_by_no.get(id_card)
            if not student_row:
                if id_card and student_no:
                    row_errors.append(f"student not found by id card/student id: {id_card}/{student_no}")
                elif id_card:
                    row_errors.append(f"student not found by id card: {id_card}")
                else:
                    row_errors.append(f"student not found by student id: {student_no}")
            else:
                student, real_school, real_grade, real_class = student_row
                if student.name != student_name:
                    row_errors.append(f"name mismatch, expected: {student.name}")
                if BatchImportService._normalize_gender(student.gender) != BatchImportService._normalize_gender(gender):
                    row_errors.append("gender mismatch")
                if real_school != school_name:
                    row_errors.append(f"school mismatch, expected: {real_school}")
                if real_grade != grade_name:
                    row_errors.append(f"grade mismatch, expected: {real_grade}")
                if real_class != class_name:
                    row_errors.append(f"class mismatch, expected: {real_class}")

            item_rules = item_name_map.get(BatchImportService._normalize_item_name_for_lookup(item_name), [])
            if item_name and not item_rules:
                item_rules = item_name_map.get(BatchImportService._normalize_item_key(item_name), [])
            selected_item = None
            if not item_rules:
                row_errors.append(f"item not in current batch standard: {item_name}")
            else:
                target_gender = BatchImportService._normalize_gender(gender)
                selected_item = next(
                    (
                        item for item in item_rules
                        if BatchImportService._normalize_gender(item.gender) in {target_gender, 'all'}
                    ),
                    None
                )
                if not selected_item:
                    if score.get("_template_style") == "wide":
                        continue
                    row_errors.append("item is not allowed for this student's gender")

            if BatchImportService._clean_cell(raw_score) == "":
                row_errors.append("score cannot be empty")
            else:
                parsed_raw_score = RuleEngine.parse_time_to_seconds(raw_score)
                if parsed_raw_score is None:
                    row_errors.append(f"invalid score format: {raw_score}")
                elif parsed_raw_score < 0 and not BatchImportService._allows_negative_score_item(selected_item):
                    row_errors.append(f"score cannot be negative: {raw_score}")

            if row_errors:
                errors.append(f"row {row_number}: {'; '.join(row_errors)}")
                continue

            normalized = dict(score)
            normalized.pop("_row_number", None)
            normalized.pop("_template_style", None)
            if student_row:
                student, real_school, real_grade, real_class = student_row
                normalized.update({
                    "student_no": student.student_no,
                    "student_name": student.name,
                    "gender": student.gender,
                    "mobile": student.phone,
                    "school_name": real_school,
                    "grade_name": real_grade,
                    "class_name": real_class
                })
            if selected_item:
                normalized.update({
                    "item_code": selected_item.item_code,
                    "item_name": selected_item.item_name
                })
            normalized_scores.append(normalized)

        return batch, normalized_scores, errors

    @staticmethod
    def parse_score_excel(file_content: bytes) -> List[Dict[str, Any]]:
        """Parse score import excel (compact or wide)."""
        try:
            wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
        except Exception:
            raise ValueError("Excel parse failed, please upload .xlsx template")
        sheet = wb.active
        scores = []

        rows = list(sheet.iter_rows())
        if len(rows) < 2:
            raise ValueError("Excel file has no importable data")

        header_map: dict[str, int] = {}
        for index, cell in enumerate(rows[0]):
            label = BatchImportService._normalize_header(cell.value)
            if label:
                header_map[label] = index
        item_header = BatchImportService._normalize_header("项目")
        legacy_item_header = BatchImportService._normalize_header("项目名称")
        if item_header not in header_map and legacy_item_header in header_map:
            header_map[item_header] = header_map[legacy_item_header]

        compact_ok = all(
            BatchImportService._normalize_header(label) in header_map
            for label in BatchImportService.SCORE_REQUIRED_LABELS
        )
        wide_ok = all(
            BatchImportService._normalize_header(label) in header_map
            for label in BatchImportService.SCORE_WIDE_BASE_REQUIRED_LABELS
        )
        template_style = 'compact' if compact_ok else 'wide' if wide_ok else None
        if template_style is None:
            raise ValueError("Template header missing required fields")

        if template_style == 'wide':
            for row_number, row in enumerate(rows[1:], start=2):
                raw_row_values = [cell.value for cell in row]
                if not any(BatchImportService._clean_cell(value) for value in raw_row_values):
                    continue

                base: dict[str, Any] = {}
                row_errors: list[str] = []
                for label, field in BatchImportService.SCORE_WIDE_HEADERS:
                    normalized_label = BatchImportService._normalize_header(label)
                    col_index = header_map.get(normalized_label)
                    raw_value = row[col_index].value if col_index is not None and col_index < len(row) else None
                    cleaned_value = BatchImportService._clean_cell(raw_value)
                    if label in BatchImportService.SCORE_WIDE_BASE_REQUIRED_LABELS and not cleaned_value:
                        row_errors.append(f"{label} cannot be blank")
                    base[field] = cleaned_value

                if row_errors:
                    raise ValueError(f"row {row_number}: {'; '.join(row_errors)}")

                entry_count = 0
                for item_label, _ in BatchImportService.SCORE_WIDE_ITEM_HEADERS:
                    candidate_labels = [item_label]
                    if item_label == "\u5750\u4f4d\u4f53\u524d\u5c48":
                        candidate_labels.append("\u5750\u4f4d\u4f53\u524d\u5f2f")
                    elif item_label == "50\u7c73\u8dd1":
                        candidate_labels.append("50\u7c73")
                    elif item_label == "1\u5206\u949f\u4ef0\u5367\u8d77\u5750":
                        candidate_labels.extend(["\u4ef0\u5367\u8d77\u5750", "\u4e00\u5206\u949f\u4ef0\u5367\u8d77\u5750"])
                    item_col = None
                    for candidate_label in candidate_labels:
                        item_col = header_map.get(BatchImportService._normalize_header(candidate_label))
                        if item_col is not None:
                            break
                    raw_value = row[item_col].value if item_col is not None and item_col < len(row) else None
                    if BatchImportService._clean_cell(raw_value) == "":
                        continue
                    scores.append({
                        **base,
                        "item_name": item_label,
                        "raw_score": raw_value,
                        "_row_number": row_number,
                        "_template_style": "wide"
                    })
                    entry_count += 1

                if entry_count == 0:
                    raise ValueError(f"row {row_number}: at least one score item is required")

            if not scores:
                raise ValueError("Excel file has no importable data")
            return scores

        for row_number, row in enumerate(rows[1:], start=2):
            raw_row_values = [cell.value for cell in row]
            if not any(BatchImportService._clean_cell(value) for value in raw_row_values):
                continue

            score: dict[str, Any] = {}
            row_errors: list[str] = []
            for label, field in BatchImportService.SCORE_HEADERS:
                col_index = header_map[BatchImportService._normalize_header(label)]
                raw_value = row[col_index].value if col_index < len(row) else None
                cleaned_value = BatchImportService._clean_cell(raw_value)
                if label in BatchImportService.SCORE_REQUIRED_LABELS and not cleaned_value:
                    row_errors.append(f"{label} cannot be blank")
                score[field] = raw_value if field == "raw_score" else cleaned_value
            if row_errors:
                raise ValueError(f"row {row_number}: {'; '.join(row_errors)}")

            score["_row_number"] = row_number
            scores.append(score)

        if not scores:
            raise ValueError("Excel file has no importable data")
        return scores

